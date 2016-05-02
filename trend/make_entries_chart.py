# -*- coding: utf-8 -*-
import sys, codecs, datetime, os, re
from openpyxl import load_workbook

def main():
    datapoints = {}
    datapoints[-1] = 0
    newface = datetime.date.today().year - 1950 # 西暦から回期を求める
    range_end = newface + 1
    for i in range(1,range_end):
        datapoints[i] = 0

    stat = os.stat(sys.argv[1])
    dt = datetime.datetime.fromtimestamp(stat.st_mtime)
    last_modified = "\nvar last_modified = \"%s\"\n" % dt.strftime("%Y-%m-%d %H:%M:%S")

    book = load_workbook(filename=sys.argv[1])
    sheet = book[u'参加者']
    maxrow = sheet.max_row
    entries = 0

    rows = sheet.iter_rows('H2:H'+str(maxrow))
    for row in rows:
        for cell in row:
            entries += 1
            val = cell.value
            name = sheet["A%d" % cell.row].value
            ticket_type = sheet["E%d" % cell.row].value
            if val in range(1,range_end):
                datapoints[val] += 1
                if val in range(newface,range_end):
                    print u'新卒: %s' % str(val),
                    print ticket_type, name
                if cell.value in range(newface-3,newface):
                    print u'学生: %s' % str(val),
                    print ticket_type, name
                if cell.value in range(newface-6,newface-3):
                    print u'学生・院生かも？: %s' % str(val),
                    print ticket_type, name
                if cell.value in range(29,30) or cell.value in range(42,43):
                    print u'当番: %s' % str(val),
                    print name
            elif val is None:
                datapoints[-1] += 1
            else:
                datapoints[-1] += 1
                print u'不正な回期入力: ',
                if not isinstance(val, unicode):
                    print str(val),name
                else:
                    print val,name

    total = "var total = %d\n" % entries

    dataPlot = "var dataPlot1 =["

    for k,v in sorted(datapoints.items(), key=lambda x: x[0]):
        if k == -1 :
            data = "{ label: %d, y: %d, indexLabel: " % (k, v)
            data += u"\"不明\" },"
        else:
            data = "{ label: %d, y: %d }," % (k, v)
        dataPlot += data

    dataPlot += "];\n"

    dataPlot += "var dataPlot2 =["

    for k,v in sorted(datapoints.items(), key=lambda x: x[1], reverse=True):
        if k == -1 :
            data = "{ label: %d, y: %d, indexLabel: " % (k, v)
            data += u"\"不明\" },"
        else:
            data = "{ label: %d, y: %d }," % (k, v)
        dataPlot += data

    dataPlot += "];"

    f = codecs.open('data.js', 'w', 'utf-8')
    f.write(dataPlot)
    f.write(last_modified)
    f.write(total)
    f.close()

if __name__ == "__main__":
    main()
